from django.contrib import admin
from django.urls import path
from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import JsonResponse
from decimal import Decimal
from .models import Brand, Tire, Disk, Supplier

import json
import uuid
import threading
import tempfile
import os


@admin.register(Supplier)
class SupplierAdmin(admin.ModelAdmin):
    list_display = ["name", "code", "is_preorder", "markup_percent", "delivery_days", "is_active", "product_count"]
    list_filter = ["is_preorder", "is_active"]
    search_fields = ["name", "code"]
    list_editable = ["markup_percent", "delivery_days", "is_active"]
    ordering = ["name"]
    actions = ["recalculate_prices", "set_markup"]

    fieldsets = (
        ("Основна інформація", {
            "fields": ("name", "code", "is_active")
        }),
        ("Налаштування наявності", {
            "fields": ("is_preorder", "delivery_days"),
            "description": "Постачальники з '21 день' в назві автоматично = 'Під замовлення'"
        }),
        ("Націнка", {
            "fields": ("markup_percent",),
            "description": "Націнка застосовується до закупівельної ціни. Після зміни натисніть 'Перерахувати ціни'"
        }),
    )

    def product_count(self, obj):
        tires = obj.tires.count()
        disks = obj.disks.count()
        return f"{tires} шин, {disks} дисків"
    product_count.short_description = "Товарів"

    @admin.action(description="Перерахувати ціни для обраних постачальників")
    def recalculate_prices(self, request, queryset):
        from .import_service import recalculate_prices_for_supplier

        total_tires = 0
        total_disks = 0

        for supplier in queryset:
            tires, disks = recalculate_prices_for_supplier(supplier)
            total_tires += tires
            total_disks += disks

        self.message_user(
            request,
            f"Ціни перераховано: {total_tires} шин, {total_disks} дисків",
            messages.SUCCESS
        )

    @admin.action(description="Встановити націнку для обраних постачальників")
    def set_markup(self, request, queryset):
        from .import_service import recalculate_prices_for_supplier

        if "apply" in request.POST:
            markup = request.POST.get("markup_percent", "").strip()
            try:
                markup = Decimal(markup)
            except Exception:
                self.message_user(request, "Невірне значення націнки", messages.ERROR)
                return

            count = queryset.update(markup_percent=markup)

            total_tires = 0
            total_disks = 0
            for supplier in queryset:
                tires, disks = recalculate_prices_for_supplier(supplier)
                total_tires += tires
                total_disks += disks

            self.message_user(
                request,
                f"Націнку {markup}% встановлено для {count} постачальників. "
                f"Перераховано: {total_tires} шин, {total_disks} дисків",
                messages.SUCCESS
            )
            return

        return render(request, "admin/set_markup.html", {
            "title": "Встановити націнку",
            "suppliers": queryset,
            "action": "set_markup",
            "opts": self.model._meta,
        })


@admin.register(Brand)
class BrandAdmin(admin.ModelAdmin):
    list_display = ["name", "slug"]
    prepopulated_fields = {"slug": ("name",)}
    search_fields = ["name"]


@admin.register(Tire)
class TireAdmin(admin.ModelAdmin):
    list_display = [
        "article",
        "brand",
        "model_name",
        "width",
        "profile",
        "diameter",
        "season",
        "studded",
        "purchase_price",
        "price",
        "in_stock",
        "is_featured",
        "supplier",
    ]
    list_filter = ["brand", "season", "studded", "in_stock", "is_featured", "diameter", "supplier"]
    search_fields = ["model_name", "article", "brand__name"]
    prepopulated_fields = {"slug": ("model_name",)}
    list_editable = ["price", "is_featured", "studded"]
    raw_id_fields = ["supplier"]


@admin.register(Disk)
class DiskAdmin(admin.ModelAdmin):
    list_display = [
        "article",
        "brand",
        "model_name",
        "diameter",
        "width",
        "bolts",
        "pcd",
        "et",
        "disk_type",
        "purchase_price",
        "price",
        "in_stock",
        "is_featured",
        "supplier",
    ]
    list_filter = ["brand", "disk_type", "in_stock", "is_featured", "diameter", "bolts", "supplier"]
    search_fields = ["model_name", "article", "brand__name"]
    prepopulated_fields = {"slug": ("model_name",)}
    list_editable = ["price", "is_featured"]
    raw_id_fields = ["supplier"]


# Custom Admin Site with import functionality
class CatalogAdminSite(admin.AdminSite):
    site_header = "КМ/Ч 120 - Адміністрування"
    site_title = "КМ/Ч 120 Admin"
    index_title = "Панель управління"

    ACTIVE_IMPORT_FILE = '/tmp/import_active_task.txt'

    def _get_active_task_id(self):
        try:
            with open(self.ACTIVE_IMPORT_FILE, 'r') as f:
                task_id = f.read().strip()
            # Verify progress file still exists (import is actually running)
            if task_id and os.path.exists(self._get_progress_file(task_id)):
                return task_id
            # Stale lock file — clean up
            try:
                os.unlink(self.ACTIVE_IMPORT_FILE)
            except OSError:
                pass
        except FileNotFoundError:
            pass
        return None

    def _set_active_task_id(self, task_id):
        with open(self.ACTIVE_IMPORT_FILE, 'w') as f:
            f.write(task_id)

    def _clear_active_task_id(self):
        try:
            os.unlink(self.ACTIVE_IMPORT_FILE)
        except OSError:
            pass

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('import-prices/', self.admin_view(self.import_prices_view), name='import_prices'),
            path('import-prices/start/', self.admin_view(self.start_import_view), name='start_import'),
            path('import-progress/<str:task_id>/', self.admin_view(self.import_progress_view), name='import_progress'),
            path('export-no-images/', self.admin_view(self.export_no_images_view), name='export_no_images'),
            path('export-skipped/<str:task_id>/', self.admin_view(self.export_skipped_view), name='export_skipped'),
            path('error-logs/', self.admin_view(self.error_logs_view), name='error_logs'),
            path('recalculate-all-prices/', self.admin_view(self.recalculate_all_prices_view), name='recalculate_all_prices'),
            path('xml-feeds/', self.admin_view(self.xml_feeds_view), name='xml_feeds'),
        ]
        return custom_urls + urls

    def _get_progress_file(self, task_id):
        return f'/tmp/import_progress_{task_id}.json'

    def _write_progress(self, task_id, data):
        progress_file = self._get_progress_file(task_id)
        tmp_file = progress_file + '.tmp'
        with open(tmp_file, 'w') as f:
            json.dump(data, f)
        os.replace(tmp_file, progress_file)

    def _run_import(self, task_id, file_path, import_type, delete_missing=False):
        import django
        django.setup()
        from .import_service import import_tires, import_disks

        def progress_callback(info):
            self._write_progress(task_id, {
                'status': 'running',
                'current': info['current'],
                'total': info['total'],
                'created': info['created'],
                'updated': info['updated'],
                'skipped': info['skipped'],
                'deleted': info.get('deleted', 0),
                'errors_count': info['errors_count'],
                'message': f"Обробка рядка {info['current']} з {info['total']}...",
            })

        try:
            if import_type == 'tires':
                result = import_tires(file_path, progress_callback=progress_callback, delete_missing=delete_missing)
            else:
                result = import_disks(file_path, progress_callback=progress_callback, delete_missing=delete_missing)

            skipped_file = ''
            skipped_rows = result.get('skipped_rows', [])
            if skipped_rows:
                skipped_file = f'/tmp/import_skipped_{task_id}.json'
                with open(skipped_file, 'w') as sf:
                    json.dump({'type': import_type, 'rows': skipped_rows}, sf, ensure_ascii=False)

            self._write_progress(task_id, {
                'status': 'completed',
                'current': result['total_rows'],
                'total': result['total_rows'],
                'created': result['created'],
                'updated': result['updated'],
                'skipped': result['skipped'],
                'deleted': result.get('deleted', 0),
                'errors_count': len(result['errors']),
                'errors': result['errors'],
                'message': 'Імпорт завершено!',
                'skipped_file': skipped_file,
            })
        except Exception as e:
            self._write_progress(task_id, {
                'status': 'error',
                'message': f'Помилка імпорту: {str(e)}',
                'current': 0,
                'total': 0,
                'created': 0,
                'updated': 0,
                'skipped': 0,
                'errors_count': 1,
                'errors': [str(e)],
            })
        finally:
            try:
                os.unlink(file_path)
            except OSError:
                pass
            self._clear_active_task_id()

    def import_prices_view(self, request):
        context = {
            'tire_count': Tire.objects.count(),
            'disk_count': Disk.objects.count(),
            'brand_count': Brand.objects.count(),
            'supplier_count': Supplier.objects.count(),
            'active_task_id': self._get_active_task_id() or '',
        }
        return render(request, 'admin/catalog/import_prices.html', context)

    def start_import_view(self, request):
        if request.method != 'POST':
            return JsonResponse({'error': 'POST only'}, status=405)

        if self._get_active_task_id():
            return JsonResponse({
                'error': 'Імпорт вже виконується. Дочекайтесь завершення.'
            }, status=409)

        import_type = request.POST.get('import_type')
        excel_file = request.FILES.get('excel_file')
        delete_missing = request.POST.get('delete_missing') == 'on'

        if not excel_file:
            return JsonResponse({'error': 'Будь ласка, виберіть файл'}, status=400)

        if import_type not in ('tires', 'disks'):
            return JsonResponse({'error': 'Невірний тип імпорту'}, status=400)

        # Save uploaded file
        suffix = '.xlsx' if excel_file.name.endswith('.xlsx') else '.xls'
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            for chunk in excel_file.chunks():
                tmp.write(chunk)
            tmp_path = tmp.name

        task_id = str(uuid.uuid4())

        # Write initial progress
        self._write_progress(task_id, {
            'status': 'running',
            'current': 0,
            'total': 0,
            'created': 0,
            'updated': 0,
            'skipped': 0,
            'deleted': 0,
            'errors_count': 0,
            'message': 'Читання файлу...',
        })

        self._set_active_task_id(task_id)

        thread = threading.Thread(
            target=self._run_import,
            args=(task_id, tmp_path, import_type, delete_missing),
            daemon=True,
        )
        thread.start()

        return JsonResponse({'task_id': task_id})

    def import_progress_view(self, request, task_id):
        progress_file = self._get_progress_file(task_id)
        try:
            with open(progress_file, 'r') as f:
                data = json.load(f)
            # Clean up progress file after client gets final status
            if data.get('status') in ('completed', 'error'):
                try:
                    os.unlink(progress_file)
                except OSError:
                    pass
                self._clear_active_task_id()
            return JsonResponse(data)
        except (FileNotFoundError, json.JSONDecodeError):
            return JsonResponse({'status': 'unknown', 'message': 'Завдання не знайдено'}, status=404)

    def recalculate_all_prices_view(self, request):
        from .import_service import recalculate_prices_for_supplier

        total_tires = 0
        total_disks = 0

        for supplier in Supplier.objects.filter(is_active=True):
            tires, disks = recalculate_prices_for_supplier(supplier)
            total_tires += tires
            total_disks += disks

        messages.success(request, f"Ціни перераховано: {total_tires} шин, {total_disks} дисків")
        return redirect('admin:import_prices')

    def export_skipped_view(self, request, task_id):
        from django.http import HttpResponse
        from openpyxl import Workbook

        skipped_file = f'/tmp/import_skipped_{task_id}.json'
        try:
            with open(skipped_file, 'r') as f:
                data = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            return HttpResponse('Файл не знайдено', status=404)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Пропущені товари'
        ws.append(['Рядок в Excel', 'Бренд', 'Модель', 'Причина'])

        for row in data.get('rows', []):
            ws.append([row['row'], row['brand'], row['model'], row['reason']])

        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 50

        import_type = data.get('type', 'items')
        count = len(data.get('rows', []))
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="skipped_{import_type}_{count}.xlsx"'
        wb.save(response)
        return response

    def export_no_images_view(self, request):
        from django.http import HttpResponse
        from django.conf import settings as conf_settings
        from openpyxl import Workbook
        from pathlib import Path

        export_type = request.GET.get('type', 'all')
        media_root = Path(conf_settings.MEDIA_ROOT)

        wb = Workbook()
        tires_count = 0
        disks_count = 0

        if export_type in ('tires', 'all'):
            ws = wb.active if export_type == 'tires' else wb.create_sheet()
            ws.title = 'Шини без картинок'
            ws.append(['Бренд', 'Модель', 'Ширина', 'Профіль', 'Діаметр', 'Сезон', 'Артикул', 'Постачальник', 'Ціна', 'Назва картинки'])

            tires = Tire.objects.select_related('brand', 'supplier').order_by('brand__name', 'model_name')
            for t in tires:
                has_file = t.image and (media_root / str(t.image)).exists()
                if not has_file:
                    ws.append([
                        t.brand.name if t.brand else '',
                        t.model_name,
                        t.width,
                        t.profile,
                        t.diameter,
                        t.get_season_display(),
                        t.article,
                        t.supplier.name if t.supplier else '',
                        float(t.price) if t.price else 0,
                        str(t.image) if t.image else '',
                    ])
                    tires_count += 1

        if export_type in ('disks', 'all'):
            if export_type == 'all':
                ws = wb.create_sheet(title='Диски без картинок')
            else:
                ws = wb.active
                ws.title = 'Диски без картинок'
            ws.append(['Бренд', 'Модель', 'Ширина', 'Діаметр', 'Болти', 'PCD', 'ET', 'Тип', 'Артикул', 'Постачальник', 'Ціна', 'Назва картинки'])

            disks = Disk.objects.select_related('brand', 'supplier').order_by('brand__name', 'model_name')
            for d in disks:
                has_file = d.image and (media_root / str(d.image)).exists()
                if not has_file:
                    ws.append([
                        d.brand.name if d.brand else '',
                        d.model_name,
                        float(d.width) if d.width else 0,
                        d.diameter,
                        d.bolts,
                        float(d.pcd) if d.pcd else 0,
                        d.et,
                        d.get_disk_type_display(),
                        d.article,
                        d.supplier.name if d.supplier else '',
                        float(d.price) if d.price else 0,
                        str(d.image) if d.image else '',
                    ])
                    disks_count += 1

        if export_type == 'all' and wb.sheetnames[0] == 'Sheet':
            del wb[wb.sheetnames[0]]

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="products_no_images_{tires_count}t_{disks_count}d.xlsx"'
        wb.save(response)
        return response

    def error_logs_view(self, request):
        from django.conf import settings

        log_file = settings.BASE_DIR / "logs" / "errors.log"
        logs = ""

        if request.method == 'POST' and request.POST.get('action') == 'clear':
            # Clear logs
            if log_file.exists():
                with open(log_file, 'w') as f:
                    f.write('')

        if log_file.exists():
            with open(log_file, 'r') as f:
                logs = f.read()
                # Show last 100 lines
                lines = logs.strip().split('\n')
                if len(lines) > 100:
                    logs = '\n'.join(lines[-100:])

        return render(request, 'admin/catalog/error_logs.html', {'logs': logs})

    def xml_feeds_view(self, request):
        base_url = request.build_absolute_uri('/').rstrip('/')
        suppliers = Supplier.objects.filter(is_active=True).order_by('name')

        return render(request, 'admin/catalog/xml_feeds.html', {
            'base_url': base_url,
            'suppliers': suppliers,
        })


# Replace default admin site
admin.site.__class__ = CatalogAdminSite
